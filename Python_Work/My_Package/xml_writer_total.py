import csv
import collections
from openpyxl import load_workbook


def writer_each_file(filepath, filename, test_item_out, test_results):
    test_result_each = []
    flag = 0
    saved_path = r"%s\%s.csv" % (filepath, filename)
    with open(saved_path, "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(test_item_out)
        for keys, values in test_results.items():
            test_result = values[:-1]
            for value in values:
                if type(value) is collections.OrderedDict:
                    for v in value.values():   # 判断P/F
                        if len(v) != 0:
                            flag += 1
                    if flag == 0:
                        test_result.append("P")
                        flag = 0
                    else:
                        test_result.append("F")
                        flag = 0
                    for v in value.values():
                        if len(v) != 0:
                            for i in range(len(v)):
                                test_result_each.append(" ".join((v[i])))
                            test_result.append(test_result_each)
                            test_result_each = []
                        else:
                            test_result.append("0")
            writer.writerow(test_result)


def writer_sum_file(test_results, test_item_out, test_flow):
    file_path = r"C:\Users\yinpeng\Desktop\Module_Test_Result\XML_Summary.xlsx"
    test_result_each = []
    flag = 0
    wb = load_workbook(file_path)
    if test_flow in wb.sheetnames:
        ws = wb[test_flow]
    else:
        ws = wb.create_sheet(test_flow)
        ws.append(test_item_out)
    for keys, values in test_results.items():
        test_result = values[:-1]
        for value in values:
            if type(value) is collections.OrderedDict:
                for v in value.values():
                    if len(v) != 0:
                        flag += 1
                if flag == 0:
                    test_result.append("P")
                    flag = 0
                else:
                    test_result.append("F")
                    flag = 0
                for v in value.values():
                    if len(v) != 0:
                        for i in range(len(v)):
                            test_result_each.append("".join((v[i])))
                        test_result.append(",".join(test_result_each))
                        test_result_each = []
                    else:
                        test_result.append("0")
        ws.append(test_result)
    wb.save(file_path)




